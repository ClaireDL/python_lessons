#!/usr/bin/env python

import openpyxl

from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook("BP3_rawdata_24549.xlsx")

print wb.sheetnames
ws1 = wb["Sheet1"]
r = ws1.max_row
c = ws1.max_column

print("Dimensions are %i rows and %i columns" % (r, c))

ws2 = wb.create_sheet()
ws2.title = "transposed"

wb.save("BP3_rawdata_24549_reformatted.xlsx")
